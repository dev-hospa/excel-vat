import openpyxl, sys


# look up for column "Ev. číslo daňového dokladu"
def find_number(sheet):
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if cell.value == "Ev. číslo daňového dokladu":
                return(cell.col_idx)

# insert values in a new sheet
def insert_filter(line, row, sheet):
    sheet.cell(row=row, column=1).value = line

def create_filter_values(sheet, column, new_sheet):
    for col in sheet.iter_cols(min_col=column, max_col=column, min_row=2):
        line = ""
        counter = 0
        row = 0
        for cell in col:
            line = line + str(cell.value) + "|" # line is a filter for Navision
            counter += 1
            if counter == 150:                  # maximum values in filter is 150
                counter = 0
                line = line.strip("|")
                row += 1
                insert_filter(line, row, new_ws)
                line = ""
            if cell == sheet.cell(sheet.max_row, column): # if you reach the last cell in column
                line = line.strip("|")
                row += 1
                insert_filter(line, row, new_ws)

                
# Open the file - the name has to be "kh.xlsx"
try:
    wb = openpyxl.load_workbook("kh.xlsx")
except:
    print("file kh.xlsx does not exist, check the name")
    sys.exit()


ws = wb.active

position = find_dic(ws)

new_wb = openpyxl.Workbook()
new_ws = new_wb.active

create_filter_values(ws, position, new_ws)

new_wb.save("filtrs.xlsx")

sys.exit()